"""
Tests for TASK-065 — Excel/XLSX export for Report Builder.

Acceptance criteria:
  1.  rows_to_xlsx returns valid .xlsx bytes + sha256.
  2.  rows_to_xlsx empty rows → valid (header-less) workbook.
  3.  rows_to_xlsx writes a bold header row.
  4.  rows_to_xlsx coerces None / UUID / Decimal safely.
  5.  build_xlsx_export task — success path writes .xlsx artifact, updates run.
  6.  build_xlsx_export task — sets artifact_format="xlsx" on the run.
  7.  build_xlsx_export task — run not found → exits cleanly (no crash).
  8.  build_xlsx_export task — run_report error captured in run.error.
  9.  definition_export POST default (no ?format) → dispatches build_csv_export.
  10. definition_export POST ?format=xlsx → dispatches build_xlsx_export.
  11. definition_export POST ?format=xlsx → run.artifact_format == "xlsx".
  12. definition_export POST ?format=invalid → 400 Bad Request.
  13. run_artifact GET for csv run → Content-Type: text/csv.
  14. run_artifact GET for xlsx run → Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.
  15. run_artifact GET for xlsx run → Content-Disposition filename ends with .xlsx.
"""

from __future__ import annotations

import uuid
from decimal import Decimal
from unittest.mock import MagicMock, patch

import pytest
from django.test import TestCase, override_settings
from rest_framework.test import APIClient

from apps.reports_builder.models import ReportDefinition, ReportRun
from apps.tenants.models import Tenant

try:
    from django.contrib.auth import get_user_model
    User = get_user_model()
except Exception:  # noqa: BLE001
    User = None

# ---------------------------------------------------------------------------
# Helpers (mirror tests_report_builder.py style)
# ---------------------------------------------------------------------------


def _make_tenant(name: str, subdomain: str) -> Tenant:
    return Tenant.objects.create(
        name=name,
        slug=subdomain,
        subdomain=subdomain,
        email=f"admin@{subdomain}.test",
        is_active=True,
    )


def _make_user(tenant, email: str, role: str = "SCHOOL_ADMIN"):
    user = User.objects.create_user(
        email=email,
        password="Pass@word1234!",
        first_name="Test",
        last_name="User",
        role=role,
    )
    user.tenant = tenant
    user.save()
    return user


def _make_definition(tenant, user, **kwargs) -> ReportDefinition:
    defaults = {
        "name": "XLSX Test Report",
        "data_source": "courses",
        "filters_json": [],
        "group_by_json": [],
        "aggregates_json": [],
    }
    defaults.update(kwargs)
    return ReportDefinition.all_objects.create(
        tenant=tenant,
        created_by=user,
        **defaults,
    )


def _make_run(tenant, definition, user, *, status="success", artifact_format="csv",
              artifact_path="/tmp/fake.csv") -> ReportRun:
    return ReportRun.all_objects.create(
        tenant=tenant,
        definition=definition,
        run_by=user,
        params_snapshot_json={
            "data_source": "courses",
            "filters": [],
            "group_by": [],
            "aggregates": [],
        },
        status=status,
        artifact_format=artifact_format,
        artifact_path=artifact_path,
        artifact_sha256="deadbeef" * 8,
    )


# ---------------------------------------------------------------------------
# Group 1 — rows_to_xlsx unit tests (no DB)
# ---------------------------------------------------------------------------


class TestRowsToXlsx(TestCase):
    """Unit tests for query_engine.rows_to_xlsx — no DB required."""

    def _import(self):
        from apps.reports_builder.query_engine import rows_to_xlsx
        return rows_to_xlsx

    def test_returns_bytes_and_sha256_hex(self):
        rows_to_xlsx = self._import()
        rows = [{"id": "1", "title": "Hello"}]
        xlsx_bytes, sha256 = rows_to_xlsx(rows)
        self.assertIsInstance(xlsx_bytes, bytes)
        self.assertGreater(len(xlsx_bytes), 0)
        self.assertIsInstance(sha256, str)
        self.assertEqual(len(sha256), 64)  # SHA-256 hex digest is 64 chars

    def test_xlsx_bytes_start_with_pk_magic(self):
        """XLSX files are ZIP archives; magic bytes are PK (0x50 0x4B)."""
        rows_to_xlsx = self._import()
        xlsx_bytes, _ = rows_to_xlsx([{"col": "value"}])
        self.assertEqual(xlsx_bytes[:2], b"PK")

    def test_empty_rows_returns_valid_xlsx(self):
        rows_to_xlsx = self._import()
        xlsx_bytes, sha256 = rows_to_xlsx([])
        self.assertIsInstance(xlsx_bytes, bytes)
        self.assertGreater(len(xlsx_bytes), 0)
        self.assertEqual(xlsx_bytes[:2], b"PK")

    def test_header_row_is_bold(self):
        """First row cells must have Font(bold=True)."""
        from openpyxl import load_workbook

        rows_to_xlsx = self._import()
        rows = [{"name": "Alice", "score": 42}]
        xlsx_bytes, _ = rows_to_xlsx(rows)

        import io
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        for cell in ws[1]:
            self.assertTrue(
                cell.font.bold,
                f"Header cell {cell.coordinate} should be bold",
            )

    def test_data_rows_written_correctly(self):
        """Data rows follow the header row with correct values."""
        from openpyxl import load_workbook
        import io

        rows_to_xlsx = self._import()
        rows = [
            {"id": "abc", "score": 99},
            {"id": "def", "score": 55},
        ]
        xlsx_bytes, _ = rows_to_xlsx(rows)

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        # Row 1 = headers; Row 2+ = data
        self.assertEqual(ws.cell(row=2, column=1).value, "abc")
        self.assertEqual(ws.cell(row=2, column=2).value, 99)
        self.assertEqual(ws.cell(row=3, column=1).value, "def")
        self.assertEqual(ws.cell(row=3, column=2).value, 55)

    def test_none_value_becomes_none_in_xlsx(self):
        """None values must be preserved (not converted to 'None' string)."""
        from openpyxl import load_workbook
        import io

        rows_to_xlsx = self._import()
        rows = [{"name": "Bob", "score": None}]
        xlsx_bytes, _ = rows_to_xlsx(rows)

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        self.assertIsNone(ws.cell(row=2, column=2).value)

    def test_uuid_coerced_to_string(self):
        """UUID values must be coerced to str (openpyxl can't write UUID objects)."""
        from openpyxl import load_workbook
        import io

        rows_to_xlsx = self._import()
        some_uuid = uuid.UUID("12345678-1234-5678-1234-567812345678")
        rows = [{"id": some_uuid, "title": "Test"}]
        xlsx_bytes, _ = rows_to_xlsx(rows)

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        cell_val = ws.cell(row=2, column=1).value
        self.assertIsInstance(cell_val, str)
        self.assertEqual(cell_val, str(some_uuid))

    def test_decimal_coerced_to_string(self):
        """Decimal values must be coerced to str (openpyxl can't write Decimal)."""
        from openpyxl import load_workbook
        import io

        rows_to_xlsx = self._import()
        rows = [{"price": Decimal("9.99")}]
        xlsx_bytes, _ = rows_to_xlsx(rows)

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        cell_val = ws.cell(row=2, column=1).value
        self.assertIsInstance(cell_val, str)
        self.assertEqual(cell_val, "9.99")

    def test_sha256_matches_bytes(self):
        """sha256_hex must be the SHA-256 of the returned bytes."""
        import hashlib

        rows_to_xlsx = self._import()
        rows = [{"x": 1}]
        xlsx_bytes, sha256 = rows_to_xlsx(rows)
        expected = hashlib.sha256(xlsx_bytes).hexdigest()
        self.assertEqual(sha256, expected)


# ---------------------------------------------------------------------------
# Group 2 — build_xlsx_export Celery task
# ---------------------------------------------------------------------------


class TestBuildXlsxExportTask(TestCase):
    """build_xlsx_export Celery task — success, format, error paths."""

    def setUp(self):
        self.tenant = _make_tenant("XLSX Task School", "xlsxtask")
        self.admin = _make_user(self.tenant, "admin@xlsxtask.test", "SCHOOL_ADMIN")
        self.definition = _make_definition(self.tenant, self.admin)

    def _run_task(self, run_id, **patch_overrides):
        defaults = {
            "apps.reports_builder.tasks.run_report": (
                [{"id": str(uuid.uuid4()), "title": "Course"}], 1
            ),
            "apps.reports_builder.tasks.rows_to_xlsx": (b"PK\x00\x00", "abc" * 21 + "a"),
        }
        defaults.update(patch_overrides)

        from apps.reports_builder.tasks import build_xlsx_export

        patchers = []
        mocks = {}
        for target, value in defaults.items():
            if callable(value) and not isinstance(value, tuple):
                p = patch(target, side_effect=value)
            elif isinstance(value, Exception):
                p = patch(target, side_effect=value)
            else:
                p = patch(target, return_value=value)
            m = p.start()
            patchers.append(p)
            mocks[target] = m

        # Patch _artifact_path to return a mock path
        apath_patcher = patch("apps.reports_builder.tasks._artifact_path")
        mock_apath = apath_patcher.start()
        patchers.append(apath_patcher)
        mock_apath.return_value.write_bytes = MagicMock()
        type(mock_apath.return_value).__str__ = lambda self: "/tmp/report.xlsx"

        try:
            build_xlsx_export(run_id)
        finally:
            for p in patchers:
                p.stop()

        return mocks

    def test_build_xlsx_export_success_sets_status_success(self):
        run = ReportRun.all_objects.create(
            tenant=self.tenant,
            definition=self.definition,
            run_by=self.admin,
            params_snapshot_json={
                "data_source": "courses",
                "filters": [],
                "group_by": [],
                "aggregates": [],
            },
            status="pending",
        )
        self._run_task(str(run.id))

        run.refresh_from_db()
        self.assertEqual(run.status, "success")

    def test_build_xlsx_export_sets_artifact_format_xlsx(self):
        """The task must set artifact_format='xlsx' on the run."""
        run = ReportRun.all_objects.create(
            tenant=self.tenant,
            definition=self.definition,
            run_by=self.admin,
            params_snapshot_json={
                "data_source": "courses",
                "filters": [],
                "group_by": [],
                "aggregates": [],
            },
            status="pending",
            artifact_format="csv",  # Default — should be overwritten
        )
        self._run_task(str(run.id))

        run.refresh_from_db()
        self.assertEqual(run.artifact_format, "xlsx")

    def test_build_xlsx_export_sets_row_count(self):
        run = ReportRun.all_objects.create(
            tenant=self.tenant,
            definition=self.definition,
            run_by=self.admin,
            params_snapshot_json={
                "data_source": "courses",
                "filters": [],
                "group_by": [],
                "aggregates": [],
            },
            status="pending",
        )
        self._run_task(str(run.id))

        run.refresh_from_db()
        self.assertEqual(run.row_count, 1)

    def test_build_xlsx_export_run_not_found_exits_cleanly(self):
        """Missing run_id must not raise — just log and return."""
        from apps.reports_builder.tasks import build_xlsx_export
        # Should not raise
        build_xlsx_export(str(uuid.uuid4()))

    def test_build_xlsx_export_run_report_error_captured(self):
        from apps.reports_builder.query_engine import ROW_CAP_EXCEEDED

        run = ReportRun.all_objects.create(
            tenant=self.tenant,
            definition=self.definition,
            run_by=self.admin,
            params_snapshot_json={
                "data_source": "courses",
                "filters": [],
                "group_by": [],
                "aggregates": [],
            },
            status="pending",
        )

        from apps.reports_builder.tasks import build_xlsx_export

        with patch(
            "apps.reports_builder.tasks.run_report",
            side_effect=ValueError(ROW_CAP_EXCEEDED),
        ):
            build_xlsx_export(str(run.id))

        run.refresh_from_db()
        self.assertEqual(run.status, "error")
        self.assertIn(ROW_CAP_EXCEEDED, run.error)

    def test_build_xlsx_export_calls_rows_to_xlsx_not_rows_to_csv(self):
        """Task must call rows_to_xlsx — not rows_to_csv — for xlsx format."""
        run = ReportRun.all_objects.create(
            tenant=self.tenant,
            definition=self.definition,
            run_by=self.admin,
            params_snapshot_json={
                "data_source": "courses",
                "filters": [],
                "group_by": [],
                "aggregates": [],
            },
            status="pending",
        )

        from apps.reports_builder.tasks import build_xlsx_export

        with patch(
            "apps.reports_builder.tasks.run_report",
            return_value=([{"id": "1"}], 1),
        ) as _, patch(
            "apps.reports_builder.tasks.rows_to_xlsx",
            return_value=(b"PK\x00", "a" * 64),
        ) as mock_xlsx, patch(
            "apps.reports_builder.tasks.rows_to_csv",
        ) as mock_csv, patch(
            "apps.reports_builder.tasks._artifact_path",
        ) as mock_apath:
            mock_apath.return_value.write_bytes = MagicMock()
            type(mock_apath.return_value).__str__ = lambda self: "/tmp/r.xlsx"
            build_xlsx_export(str(run.id))

        mock_xlsx.assert_called_once()
        mock_csv.assert_not_called()


# ---------------------------------------------------------------------------
# Group 3 — definition_export view with ?format=xlsx
# ---------------------------------------------------------------------------


class TestDefinitionExportFormatParam(TestCase):
    """definition_export POST — ?format=csv (default) and ?format=xlsx routing."""

    def setUp(self):
        self.tenant = _make_tenant("Export Format School", "exportfmt")
        self.admin = _make_user(self.tenant, "admin@exportfmt.test", "SCHOOL_ADMIN")
        self.definition = _make_definition(self.tenant, self.admin)

    def _post_export(self, format_param=None, extra_qs=""):
        client = APIClient()
        client.force_authenticate(user=self.admin)
        url = f"/api/v1/admin/reports/definitions/{self.definition.id}/export/"
        if format_param is not None:
            url += f"?format={format_param}"
        elif extra_qs:
            url += f"?{extra_qs}"

        with patch("utils.tenant_middleware.get_current_tenant", return_value=self.tenant), \
             patch("apps.reports_builder.views.cache.get", return_value=0), \
             patch("apps.reports_builder.views.cache.set"), \
             patch("apps.reports_builder.tasks.build_csv_export.delay") as mock_csv, \
             patch("apps.reports_builder.tasks.build_xlsx_export.delay") as mock_xlsx:
            resp = client.post(
                url,
                format="json",
                HTTP_HOST=f"{self.tenant.subdomain}.localhost",
            )
            return resp, mock_csv, mock_xlsx

    def test_default_no_format_dispatches_csv_task(self):
        """No ?format param → build_csv_export is dispatched."""
        resp, mock_csv, mock_xlsx = self._post_export()
        self.assertEqual(resp.status_code, 202)
        mock_csv.assert_called_once()
        mock_xlsx.assert_not_called()

    def test_format_csv_dispatches_csv_task(self):
        """?format=csv → build_csv_export is dispatched."""
        resp, mock_csv, mock_xlsx = self._post_export(format_param="csv")
        self.assertEqual(resp.status_code, 202)
        mock_csv.assert_called_once()
        mock_xlsx.assert_not_called()

    def test_format_xlsx_dispatches_xlsx_task(self):
        """?format=xlsx → build_xlsx_export is dispatched."""
        resp, mock_csv, mock_xlsx = self._post_export(format_param="xlsx")
        self.assertEqual(resp.status_code, 202)
        mock_xlsx.assert_called_once()
        mock_csv.assert_not_called()

    def test_format_xlsx_sets_artifact_format_on_run(self):
        """?format=xlsx → ReportRun.artifact_format is 'xlsx'."""
        resp, mock_csv, mock_xlsx = self._post_export(format_param="xlsx")
        self.assertEqual(resp.status_code, 202)
        run_id = resp.json()["run_id"]
        run = ReportRun.all_objects.get(id=run_id)
        self.assertEqual(run.artifact_format, "xlsx")

    def test_format_csv_sets_artifact_format_csv_on_run(self):
        """?format=csv → ReportRun.artifact_format is 'csv'."""
        resp, mock_csv, mock_xlsx = self._post_export(format_param="csv")
        self.assertEqual(resp.status_code, 202)
        run_id = resp.json()["run_id"]
        run = ReportRun.all_objects.get(id=run_id)
        self.assertEqual(run.artifact_format, "csv")

    def test_invalid_format_returns_400(self):
        """?format=pdf → 400 Bad Request (only csv and xlsx are supported)."""
        resp, mock_csv, mock_xlsx = self._post_export(format_param="pdf")
        self.assertEqual(resp.status_code, 400)
        self.assertIn("format", str(resp.json()).lower())


# ---------------------------------------------------------------------------
# Group 4 — run_artifact view Content-Type routing
# ---------------------------------------------------------------------------


class TestRunArtifactContentType(TestCase):
    """run_artifact GET — Content-Type driven by run.artifact_format."""

    def setUp(self):
        self.tenant = _make_tenant("Artifact CT School", "artifactct")
        self.admin = _make_user(self.tenant, "admin@artifactct.test", "SCHOOL_ADMIN")
        self.definition = _make_definition(self.tenant, self.admin)

    def _get_artifact(self, run, file_bytes=b"data"):
        from apps.courses.helpers.signed_urls import make_signed_url

        client = APIClient()
        client.force_authenticate(user=self.admin)

        base_url = f"http://testserver/api/v1/admin/reports/runs/{run.id}/artifact/"
        signed = make_signed_url(
            base_url=base_url,
            user_id=str(self.admin.id),
            ttl_seconds=3600,
            extra_params={"run": str(run.id)},
        )
        from urllib.parse import parse_qs, urlparse
        qs = parse_qs(urlparse(signed).query)
        token = qs["lp_token"][0]
        expires = int(qs["lp_expires"][0])

        url = (
            f"/api/v1/admin/reports/runs/{run.id}/artifact/"
            f"?lp_token={token}&lp_expires={expires}&run={run.id}"
        )

        with patch("utils.tenant_middleware.get_current_tenant", return_value=self.tenant), \
             patch("builtins.open", MagicMock(
                 return_value=MagicMock(
                     __enter__=MagicMock(return_value=MagicMock(read=MagicMock(return_value=file_bytes))),
                     __exit__=MagicMock(return_value=False),
                 )
             )):
            return client.get(
                url,
                HTTP_HOST=f"{self.tenant.subdomain}.localhost",
            )

    def test_csv_run_artifact_content_type_is_text_csv(self):
        """A run with artifact_format='csv' must be served as text/csv."""
        run = _make_run(
            self.tenant, self.definition, self.admin,
            artifact_format="csv",
            artifact_path="/tmp/fake.csv",
        )
        resp = self._get_artifact(run, file_bytes=b"id,title\n1,Course\n")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("text/csv", resp.get("Content-Type", ""))

    def test_xlsx_run_artifact_content_type_is_xlsx(self):
        """A run with artifact_format='xlsx' must be served as the OOXML MIME type."""
        run = _make_run(
            self.tenant, self.definition, self.admin,
            artifact_format="xlsx",
            artifact_path="/tmp/fake.xlsx",
        )
        resp = self._get_artifact(run, file_bytes=b"PK\x03\x04fake-xlsx")
        self.assertEqual(resp.status_code, 200)
        expected_ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        self.assertIn(expected_ct, resp.get("Content-Type", ""))

    def test_xlsx_run_artifact_content_disposition_has_xlsx_extension(self):
        """Content-Disposition filename must end in .xlsx for xlsx runs."""
        run = _make_run(
            self.tenant, self.definition, self.admin,
            artifact_format="xlsx",
            artifact_path="/tmp/fake.xlsx",
        )
        resp = self._get_artifact(run, file_bytes=b"PK\x03\x04fake-xlsx")
        self.assertEqual(resp.status_code, 200)
        cd = resp.get("Content-Disposition", "")
        self.assertIn(".xlsx", cd)

    def test_csv_run_artifact_content_disposition_has_csv_extension(self):
        """Content-Disposition filename must end in .csv for csv runs."""
        run = _make_run(
            self.tenant, self.definition, self.admin,
            artifact_format="csv",
            artifact_path="/tmp/fake.csv",
        )
        resp = self._get_artifact(run, file_bytes=b"id,title\n")
        self.assertEqual(resp.status_code, 200)
        cd = resp.get("Content-Disposition", "")
        self.assertIn(".csv", cd)
